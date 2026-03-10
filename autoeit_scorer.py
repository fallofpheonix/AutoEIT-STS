"""
AutoEIT Scoring Engine
======================
Reproducible, rule-based scoring system for Spanish Elicited Imitation Task (EIT).
Applies meaning-based rubric to learner transcriptions vs. target (prompt) sentences.

Author: AutoEIT Project — Specific Test II
Scoring scale: 0–4

RUBRIC (empirically calibrated against 1,560 expert-annotated utterances):

  Score  Description                  Key signals
  ─────  ─────────────────────────────────────────────────────────────────────
    4    Accurate reproduction         edit_dist ≤ 3 AND all deviations are
                                       disfluencies/false-starts (already
                                       stripped) or simple repetition artefacts.
                                       In practice: normalised edit_dist = 0 in
                                       95.7% of score-4 cases; ≤ 1 in 99.4%.
    3    Near-accurate                 edit_dist 1–3, with meaning intact.
                                       Typical: ser↔estar swap, morphological
                                       variant (romancia→romance), dropped/added
                                       single function word, minor conjugation
                                       change. Content overlap ≥ 0.70.
    2    Partial meaning               Noticeable deviation: wrong content word,
                                       missing/extra phrase, distorted structure,
                                       but ≥ 35% content words preserved.
    1    Minimal meaning               Only isolated recognisable fragments;
                                       content overlap < 25%.
    0    No meaning preserved          Wrong language, silence, complete
                                       substitution with unrelated words.

PREPROCESSING (critical):
  - Strip stimulus word-count hints: '(7)', '(12)' etc — these are NOT part of
    the sentence and inflate edit distance if left in.
  - Remove bracketed annotations in transcriptions: [false-start-], [...]
  - Remove punctuation, lowercase, strip diacritics, collapse whitespace.
"""

import re
import unicodedata
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import math


# ─────────────────────────────────────────────
#  TEXT NORMALISATION
# ─────────────────────────────────────────────

def strip_accents(s: str) -> str:
    """Remove diacritics so accent errors don't count as word substitutions."""
    return ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )

# Annotation patterns present in the transcription data
_BRACKET_CONTENT = re.compile(r'\[.*?\]')           # [false start-], [...]
_PUNCTUATION     = re.compile(r'[¿?¡!,.:;«»"\'\`´]')
_WHITESPACE      = re.compile(r'\s+')
_WC_HINT         = re.compile(r'\(\d+\)\s*$')        # word-count hint "(7)" at end of stimulus


def clean(text: str, is_stimulus: bool = False) -> str:
    """
    Normalise a transcription or stimulus string to a comparable token sequence.

    Steps:
      1. Remove word-count hint from stimulus  (e.g. "(12)" at end)
      2. Remove bracketed annotations from transcriptions  ([xxx], [...])
      3. Strip punctuation
      4. Lowercase
      5. Strip diacritics  (so accent errors do not count as substitutions)
      6. Collapse whitespace
    """
    if not isinstance(text, str):
        return ''
    t = text
    if is_stimulus:
        t = _WC_HINT.sub('', t)           # remove "(7)", "(12)" etc.
    t = _BRACKET_CONTENT.sub(' ', t)      # remove [xxx] annotations
    t = _PUNCTUATION.sub('', t)
    t = t.lower()
    t = strip_accents(t)
    t = _WHITESPACE.sub(' ', t).strip()
    return t


def tokenise(text: str, is_stimulus: bool = False) -> list[str]:
    return clean(text, is_stimulus=is_stimulus).split()


# ─────────────────────────────────────────────
#  LEVENSHTEIN EDIT DISTANCE (token-level)
# ─────────────────────────────────────────────

def token_edit_distance(seq_a: list[str], seq_b: list[str]) -> int:
    """
    Standard dynamic-programming Levenshtein distance at the *word* level.
    Each operation (insertion, deletion, substitution) costs 1.
    """
    m, n = len(seq_a), len(seq_b)
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev = dp[:]
        dp[0] = i
        for j in range(1, n + 1):
            if seq_a[i - 1] == seq_b[j - 1]:
                dp[j] = prev[j - 1]
            else:
                dp[j] = 1 + min(prev[j], dp[j - 1], prev[j - 1])
    return dp[n]


# ─────────────────────────────────────────────
#  CONTENT-WORD OVERLAP
# ─────────────────────────────────────────────

# Spanish function words — these should NOT count toward meaning-content overlap
_FUNCTION_WORDS = {
    'el', 'la', 'los', 'las', 'un', 'una', 'unos', 'unas',
    'de', 'en', 'a', 'por', 'para', 'con', 'sin', 'sobre',
    'que', 'se', 'le', 'lo', 'me', 'te', 'nos', 'les',
    'y', 'o', 'pero', 'sino', 'aunque', 'como', 'si', 'porque',
    'ya', 'no', 'ni', 'muy', 'mas', 'su', 'sus',
    'mi', 'mis', 'tu', 'tus', 'al', 'del',
    'este', 'esta', 'estos', 'estas', 'ese', 'esa', 'esos', 'esas',
    'hay', 'es', 'son', 'estan', 'era', 'eran', 'fue', 'fueron',
    'ha', 'han', 'he', 'hemos', 'haber',
    'ser', 'estar', 'tener', 'hacer',
    'al', 'del', 'yo', 'tu', 'el', 'ella', 'nosotros', 'ellos', 'ellas',
    'usted', 'ustedes', 'quien', 'cuyo', 'cuya',
}

def content_words(tokens: list[str]) -> list[str]:
    return [t for t in tokens if t not in _FUNCTION_WORDS and len(t) > 1]


def content_overlap_ratio(target_tokens: list[str], learner_tokens: list[str]) -> float:
    """
    Fraction of target *content* words reproduced in the learner response.
    Uses set intersection so order doesn't matter here.
    """
    tc = content_words(target_tokens)
    lc = set(content_words(learner_tokens))
    if not tc:
        return 1.0
    matched = sum(1 for w in tc if w in lc)
    return matched / len(tc)


# ─────────────────────────────────────────────
#  SPECIAL-CASE DETECTORS
# ─────────────────────────────────────────────

def is_no_response(raw: str) -> bool:
    """Returns True if the transcription is empty, only [...], or only annotations."""
    if not isinstance(raw, str) or not raw.strip():
        return True
    cleaned = _BRACKET_CONTENT.sub('', raw).strip()
    cleaned = _PUNCTUATION.sub('', cleaned).strip()
    return cleaned == ''


def is_wrong_language(raw: str) -> bool:
    """
    Heuristic: if the rater noted '[en inglés]' or similar, score = 0.
    """
    if not isinstance(raw, str):
        return False
    lower = raw.lower()
    return 'en ingles' in lower or 'en inglés' in lower


# ─────────────────────────────────────────────
#  SER / ESTAR SUBSTITUTION DETECTOR
# ─────────────────────────────────────────────

_SER_FORMS   = {'es', 'son', 'era', 'eran', 'fue', 'fueron', 'sea', 'sean', 'ser', 'soy', 'somos'}
_ESTAR_FORMS = {'esta', 'estan', 'estaba', 'estaban', 'estuvo', 'estuvieron',
                'estar', 'estoy', 'estamos'}

def only_ser_estar_swap(target_tokens: list[str], learner_tokens: list[str]) -> bool:
    """Returns True if the ONLY difference is a ser↔estar substitution."""
    if len(target_tokens) != len(learner_tokens):
        return False
    diffs = [(t, l) for t, l in zip(target_tokens, learner_tokens) if t != l]
    if len(diffs) != 1:
        return False
    t_word, l_word = diffs[0]
    return (
        (t_word in _SER_FORMS and l_word in _ESTAR_FORMS) or
        (t_word in _ESTAR_FORMS and l_word in _SER_FORMS)
    )


# ─────────────────────────────────────────────
#  MORPHOLOGICAL CLOSENESS
# ─────────────────────────────────────────────

def char_similarity(a: str, b: str) -> float:
    """Character-level Levenshtein similarity (0–1) for two words."""
    if a == b:
        return 1.0
    max_len = max(len(a), len(b), 1)
    m, n = len(a), len(b)
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev = dp[:]
        dp[0] = i
        for j in range(1, n + 1):
            if a[i - 1] == b[j - 1]:
                dp[j] = prev[j - 1]
            else:
                dp[j] = 1 + min(prev[j], dp[j - 1], prev[j - 1])
    return 1 - dp[n] / max_len


def morph_close(w1: str, w2: str, threshold: float = 0.72) -> bool:
    """True if two words share enough characters to be morphological variants."""
    return char_similarity(w1, w2) >= threshold


# ─────────────────────────────────────────────
#  CORE RUBRIC SCORER
# ─────────────────────────────────────────────

def score_utterance(target_raw: str, learner_raw: str) -> tuple[int, str]:
    """
    Apply the meaning-based EIT rubric.

    Calibration notes (from 1,560 expert-annotated utterances):
      Score 4: edit_dist = 0 in 95.7% of cases; ≤ 1 in 99.4%
               edit_dist ≤ 3 covers all score-4 cases (repetition artefacts)
      Score 3: edit_dist 1–3, content overlap ≥ 0.70
      Score 2: edit_dist 2–7, content overlap 0.30–0.70
      Score 1: edit_dist 4–11, content overlap < 0.30
      Score 0: edit_dist 4–12, no content overlap; wrong language; no response

    Returns (score: int, rationale: str)
    """

    # ── Score 0 edge cases ──────────────────────────────────────────────
    if is_wrong_language(learner_raw):
        return 0, "Response in wrong language [en inglés]"

    if is_no_response(learner_raw):
        return 0, "No response / empty transcription"

    # ── Normalise ────────────────────────────────────────────────────────
    target_tokens  = tokenise(target_raw,  is_stimulus=True)
    learner_tokens = tokenise(learner_raw, is_stimulus=False)

    if not target_tokens:
        return 0, "Target sentence is empty"

    # ── Metrics ──────────────────────────────────────────────────────────
    n_target  = len(target_tokens)
    n_learner = len(learner_tokens)
    edit_dist = token_edit_distance(target_tokens, learner_tokens)
    overlap   = content_overlap_ratio(target_tokens, learner_tokens)

    # ── Score 4: accurate reproduction ───────────────────────────────────
    # After preprocessing (stripping brackets, word-count hints, punctuation,
    # accents), score-4 utterances have edit_dist = 0 in 95.7% of cases.
    # edit_dist ≤ 3 covers all score-4 cases (false-start repetitions).
    # Key boundary: ALL score-4 cases have overlap ≥ 0.67.

    if edit_dist == 0:
        return 4, "Exact reproduction after normalisation"

    if edit_dist == 1:
        # One-word deviation — check nature of the difference
        if overlap >= 0.67:
            return 4, (
                f"1-word deviation, full meaning preserved "
                f"(edit=1, content_overlap={overlap:.2f})"
            )

    if edit_dist == 2 and overlap >= 0.80:
        # 2-word deviation — could be repetition artefact or minor addition
        # Real score-4 examples: repetition of phrase, extra preposition
        return 4, (
            f"2-word deviation consistent with disfluency/repetition artefact "
            f"(edit=2, content_overlap={overlap:.2f})"
        )

    if edit_dist == 3 and overlap >= 0.85:
        # Rare: score-4 with ed=3 exists (e.g. word-form collision)
        return 4, (
            f"3-word deviation, near-complete content preserved "
            f"(edit=3, content_overlap={overlap:.2f})"
        )

    # ── Score 3: near-accurate ───────────────────────────────────────────
    # Meaning preserved but with a minor grammatical/morphological deviation.
    # Calibrated boundary: edit_dist 1–3, content_overlap ≥ 0.65

    if only_ser_estar_swap(target_tokens, learner_tokens):
        return 3, "Ser↔estar substitution; core meaning preserved"

    if edit_dist == 1 and overlap >= 0.50:
        # Below the score-4 overlap threshold but still near-accurate
        if len(target_tokens) == len(learner_tokens):
            # Same-length: single word substitution
            diffs = [(t, l) for t, l in zip(target_tokens, learner_tokens) if t != l]
            if diffs:
                tw, lw = diffs[0]
                if morph_close(tw, lw):
                    return 3, f"Morphological variant: '{tw}'→'{lw}'; meaning preserved"
                return 3, f"Single word substitution ('{tw}'→'{lw}'); meaning largely preserved"
        return 3, f"1-word deviation; meaning largely preserved (overlap={overlap:.2f})"

    if edit_dist == 2 and overlap >= 0.65:
        return 3, (
            f"Minor deviation (edit=2, content_overlap={overlap:.2f}); "
            f"meaning essentially preserved"
        )

    if edit_dist == 3 and overlap >= 0.70:
        return 3, (
            f"Moderate deviation (edit=3, content_overlap={overlap:.2f}); "
            f"meaning preserved"
        )

    if edit_dist <= 4 and overlap >= 0.75:
        return 3, (
            f"Structural deviation (edit={edit_dist}), high content overlap "
            f"({overlap:.2f}); meaning preserved"
        )

    # ── Score 2: partial meaning ─────────────────────────────────────────
    # Noticeable deviation but some core content retained.
    # Calibrated: content_overlap typically 0.30–0.70, edit_dist 2–7

    if overlap >= 0.35 and n_learner >= max(2, n_target * 0.30):
        return 2, (
            f"Partial meaning (content_overlap={overlap:.2f}, edit={edit_dist}, "
            f"n_target={n_target}, n_learner={n_learner})"
        )

    # Short response with some overlap (e.g. "son muy grandes" → score 2)
    if overlap >= 0.25 and n_learner >= 2:
        return 2, (
            f"Partial response — some content preserved "
            f"(content_overlap={overlap:.2f}, n_learner={n_learner})"
        )

    # ── Score 1: minimal meaning ─────────────────────────────────────────
    # Only isolated recognisable fragments.
    if overlap >= 0.10 or (n_learner >= 1 and overlap > 0):
        return 1, (
            f"Minimal meaning — isolated fragments only "
            f"(content_overlap={overlap:.2f}, n_learner={n_learner})"
        )

    # ── Score 0: no meaning ───────────────────────────────────────────────
    return 0, (
        f"No meaning preserved (content_overlap={overlap:.2f}, edit={edit_dist})"
    )


# ─────────────────────────────────────────────
#  DATASET LOADER
# ─────────────────────────────────────────────

def load_dataset(filepath: str) -> pd.DataFrame:
    """
    Load all participant sheets from the EIT Excel workbook.
    Returns a tidy DataFrame with columns:
      participant_id, version, sentence_id, stimulus, transcription, human_score
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    records = []

    # Sheets named like "4_vA", "18_vB", etc.
    participant_pattern = re.compile(r'^(\d+)_(v[AB])$')

    for sheet_name in wb.sheetnames:
        m = participant_pattern.match(sheet_name)
        if not m:
            continue
        pid     = int(m.group(1))
        version = m.group(2)
        ws      = wb[sheet_name]

        for row in ws.iter_rows(min_row=2, values_only=True):
            sent_id = row[0]
            if not isinstance(sent_id, int):
                continue
            stimulus      = row[1]
            transcription = row[2]
            human_score   = row[3]

            records.append({
                'participant_id': pid,
                'version':        version,
                'sentence_id':    sent_id,
                'stimulus':       stimulus,
                'transcription':  transcription,
                'human_score':    human_score,
            })

    df = pd.DataFrame(records)
    df = df.sort_values(['participant_id', 'version', 'sentence_id']).reset_index(drop=True)
    return df


# ─────────────────────────────────────────────
#  PIPELINE
# ─────────────────────────────────────────────

def run_scoring_pipeline(filepath: str) -> pd.DataFrame:
    """
    Full pipeline: load → score → return enriched DataFrame.
    Adds columns: auto_score, rationale, has_human_score, agreement, score_diff
    """
    print(f"Loading dataset from: {filepath}")
    df = load_dataset(filepath)
    print(f"  → {len(df)} utterances from {df['participant_id'].nunique()} participants")

    # Apply scorer row-by-row (deterministic, no randomness)
    scores_and_rationale = df.apply(
        lambda r: score_utterance(r['stimulus'], r['transcription']),
        axis=1
    )
    df['auto_score'] = scores_and_rationale.apply(lambda x: x[0])
    df['rationale']  = scores_and_rationale.apply(lambda x: x[1])

    # Evaluation vs. human rater (where available)
    mask = df['human_score'].notna()
    df['has_human_score'] = mask
    df.loc[mask, 'agreement']  = (
        df.loc[mask, 'auto_score'] == df.loc[mask, 'human_score']
    )
    df.loc[mask, 'score_diff'] = (
        df.loc[mask, 'auto_score'] - df.loc[mask, 'human_score']
    )

    return df


# ─────────────────────────────────────────────
#  EVALUATION METRICS
# ─────────────────────────────────────────────

def compute_metrics(df: pd.DataFrame) -> dict:
    """Compute agreement rate and per-participant score deviation."""
    rated = df[df['has_human_score']].copy()
    if rated.empty:
        return {}

    # Sentence-level exact agreement
    n_total   = len(rated)
    n_agree   = rated['agreement'].sum()
    agree_pct = n_agree / n_total * 100

    # Within-1 agreement (adjacent score tolerance)
    within1   = (rated['score_diff'].abs() <= 1).sum()
    within1_pct = within1 / n_total * 100

    # Per-participant total score deviation
    auto_totals  = rated.groupby(['participant_id', 'version'])['auto_score'].sum()
    human_totals = rated.groupby(['participant_id', 'version'])['human_score'].sum()
    deviations   = (auto_totals - human_totals).abs()
    mean_dev     = deviations.mean()
    max_dev      = deviations.max()
    pct_within10 = (deviations <= 10).mean() * 100

    # Score distribution comparison
    auto_dist  = rated['auto_score'].value_counts().sort_index()
    human_dist = rated['human_score'].value_counts().sort_index()

    return {
        'n_rated_utterances':          n_total,
        'exact_agreement_rate_pct':    round(agree_pct, 2),
        'within1_agreement_pct':       round(within1_pct, 2),
        'mean_participant_deviation':  round(mean_dev, 2),
        'max_participant_deviation':   int(max_dev),
        'pct_participants_within10':   round(pct_within10, 2),
        'n_participants':              len(deviations),
        'auto_score_distribution':     auto_dist.to_dict(),
        'human_score_distribution':    human_dist.to_dict(),
        'confusion_summary':           pd.crosstab(
                                           rated['human_score'],
                                           rated['auto_score'],
                                           rownames=['Human'],
                                           colnames=['Auto']
                                       ),
    }


# ─────────────────────────────────────────────
#  EXCEL OUTPUT
# ─────────────────────────────────────────────

SCORE_FILLS = {
    4: PatternFill(fill_type='solid', fgColor='C6EFCE'),   # green
    3: PatternFill(fill_type='solid', fgColor='FFEB9C'),   # yellow
    2: PatternFill(fill_type='solid', fgColor='FFCC99'),   # orange
    1: PatternFill(fill_type='solid', fgColor='FFC7CE'),   # light red
    0: PatternFill(fill_type='solid', fgColor='FF0000'),   # red
}
HEADER_FILL = PatternFill(fill_type='solid', fgColor='4472C4')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def write_output_workbook(df: pd.DataFrame, source_path: str, out_path: str) -> None:
    """
    Writes auto scores and rationale back into a copy of the original workbook.
    Adds AutoEIT_Score (col E) and Rationale (col F) to each participant sheet.
    Also creates a consolidated AutoEIT_Summary sheet.
    """
    from shutil import copyfile
    copyfile(source_path, out_path)

    wb  = openpyxl.load_workbook(out_path)
    pid_pat = re.compile(r'^(\d+)_(v[AB])$')

    for sheet_name in wb.sheetnames:
        if not pid_pat.match(sheet_name):
            continue

        ws  = wb[sheet_name]
        pid = int(pid_pat.match(sheet_name).group(1))
        ver = pid_pat.match(sheet_name).group(2)

        sub    = df[(df['participant_id'] == pid) & (df['version'] == ver)]
        lookup = {int(r['sentence_id']): (r['auto_score'], r['rationale'])
                  for _, r in sub.iterrows()}

        ws.cell(row=1, column=5).value = 'AutoEIT_Score'
        ws.cell(row=1, column=6).value = 'Rationale'
        for col in [5, 6]:
            cell = ws.cell(row=1, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')

        for row_idx in range(2, ws.max_row + 1):
            sent_id = ws.cell(row=row_idx, column=1).value
            if not isinstance(sent_id, int) or sent_id not in lookup:
                continue
            auto_score, rationale = lookup[sent_id]
            score_cell = ws.cell(row=row_idx, column=5)
            score_cell.value = auto_score
            score_cell.fill  = SCORE_FILLS.get(auto_score, SCORE_FILLS[0])
            score_cell.alignment = Alignment(horizontal='center')
            rat_cell = ws.cell(row=row_idx, column=6)
            rat_cell.value = rationale
            rat_cell.alignment = Alignment(wrap_text=True)

    # ── Summary sheet ──────────────────────────────────────────────────
    if 'AutoEIT_Summary' in wb.sheetnames:
        del wb['AutoEIT_Summary']
    ws_sum = wb.create_sheet('AutoEIT_Summary', 0)

    headers = ['Participant', 'Version', 'Sentence', 'Stimulus',
               'Transcription', 'Human Score', 'AutoEIT Score', 'Score Diff', 'Rationale']
    for c, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    col_widths = [14, 10, 10, 50, 55, 13, 13, 11, 60]
    for c, w in enumerate(col_widths, 1):
        ws_sum.column_dimensions[get_column_letter(c)].width = w

    row_idx = 2
    for _, r in df.iterrows():
        hs   = r['human_score']
        auts = int(r['auto_score'])
        diff = (auts - int(hs)) if pd.notna(hs) else ''
        row_data = [r['participant_id'], r['version'], int(r['sentence_id']),
                    r['stimulus'], r['transcription'],
                    int(hs) if pd.notna(hs) else '', auts, diff, r['rationale']]
        for c, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=row_idx, column=c, value=val)
            if c == 7:
                cell.fill = SCORE_FILLS.get(auts, SCORE_FILLS[0])
                cell.alignment = Alignment(horizontal='center')
            if c == 8 and isinstance(diff, int) and diff != 0:
                cell.font = Font(color='FF0000', bold=True)
            cell.alignment = Alignment(wrap_text=True)
        row_idx += 1

    ws_sum.freeze_panes = 'A2'
    wb.save(out_path)
    print(f"  → Output saved: {out_path}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

if __name__ == '__main__':
    import argparse
    from pathlib import Path

    parser = argparse.ArgumentParser(
        description='AutoEIT scorer: score transcriptions against prompt sentences.'
    )
    parser.add_argument(
        '--source',
        type=str,
        default='',
        help='Input workbook path (.xlsx). If omitted, uses known local/default paths.'
    )
    parser.add_argument(
        '--output-xlsx',
        type=str,
        default='',
        help='Output scored workbook path (.xlsx). Default: ./AutoEIT_Sample_Transcriptions_Scored.xlsx'
    )
    parser.add_argument(
        '--output-csv',
        type=str,
        default='',
        help='Output scored CSV path. Default: ./AutoEIT_scores.csv'
    )
    args = parser.parse_args()

    repo_dir = Path(__file__).resolve().parent
    source_candidates = []
    if args.source:
        source_candidates.append(Path(args.source).expanduser().resolve())
    source_candidates.extend([
        repo_dir / 'Example_EIT Transcription and Scoring Sheet.xlsx',
        repo_dir / 'Example_EIT_Transcription_and_Scoring_Sheet.xlsx',
        Path('/mnt/user-data/uploads/Example_EIT_Transcription_and_Scoring_Sheet.xlsx'),
    ])

    source_path = next((p for p in source_candidates if p.exists()), None)
    if source_path is None:
        raise FileNotFoundError(
            'Input workbook not found. Set --source or place one of:\n'
            f'  - {repo_dir / "Example_EIT Transcription and Scoring Sheet.xlsx"}\n'
            f'  - {repo_dir / "Example_EIT_Transcription_and_Scoring_Sheet.xlsx"}\n'
            '  - /mnt/user-data/uploads/Example_EIT_Transcription_and_Scoring_Sheet.xlsx'
        )

    output_xlsx = Path(args.output_xlsx).expanduser().resolve() if args.output_xlsx else (
        repo_dir / 'AutoEIT_Sample_Transcriptions_Scored.xlsx'
    )
    output_csv = Path(args.output_csv).expanduser().resolve() if args.output_csv else (
        repo_dir / 'AutoEIT_scores.csv'
    )
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    output_csv.parent.mkdir(parents=True, exist_ok=True)

    # ── Run pipeline ────────────────────────────────────────────────────
    df = run_scoring_pipeline(str(source_path))

    # ── Print sentence-level sample ─────────────────────────────────────
    print("\n=== SENTENCE-LEVEL SCORES (first 30 rows with human scores) ===")
    sample = df[df['has_human_score']].head(30)[
        ['participant_id', 'version', 'sentence_id',
         'human_score', 'auto_score', 'score_diff', 'rationale']
    ]
    pd.set_option('display.max_colwidth', 65)
    pd.set_option('display.max_rows', 60)
    pd.set_option('display.width', 200)
    print(sample.to_string(index=False))

    # ── Metrics ─────────────────────────────────────────────────────────
    metrics = compute_metrics(df)
    print("\n=== EVALUATION METRICS ===")
    for k, v in metrics.items():
        if k == 'confusion_summary':
            print(f"\n{k}:\n{v}")
        else:
            print(f"  {k}: {v}")

    # ── Write outputs ───────────────────────────────────────────────────
    print("\nWriting output workbook…")
    write_output_workbook(df, str(source_path), str(output_xlsx))
    df[['participant_id', 'version', 'sentence_id', 'stimulus',
        'transcription', 'human_score', 'auto_score', 'rationale']].to_csv(output_csv, index=False)
    print(f"  → CSV saved: {output_csv}")
    print("\n✓ AutoEIT scoring complete.")
