"""Tests for src.io — workbook parsing, schema validation, and writing."""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from src.io import (
    ensure_parent_dir,
    last_populated_header_column,
    load_dataset,
    parse_sheet_identity,
    validate_workbook_schema,
    write_csv_outputs,
    write_output_workbook,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_workbook(sheets: dict[str, list[tuple]]) -> Path:
    """Write a temporary workbook with the given sheets and return its path."""
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = Path(tmp.name)
    tmp.close()  # close handle before openpyxl writes to the same path
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    wb.save(tmp_path)
    return tmp_path


_VALID_SHEET = [
    ("Sentence", "Stimulus", "Transcription", "Human Score"),
    (1, "El gato corre", "el gato corre", 4),
    (2, "La niña lee", "la nina lee", 3),
]

_VALID_SHEET_NO_HUMAN = [
    ("Sentence", "Stimulus", "Transcription"),
    (1, "El gato corre", "el gato corre"),
]


# ---------------------------------------------------------------------------
# parse_sheet_identity
# ---------------------------------------------------------------------------


class TestParseSheetIdentity:
    def test_numeric_underscore_vA(self):
        pid, version = parse_sheet_identity("42_vA")
        assert pid == "42"
        assert version == "vA"

    def test_numeric_underscore_vB(self):
        pid, version = parse_sheet_identity("7_vB")
        assert pid == "7"
        assert version == "vB"

    def test_numeric_dash_1A(self):
        pid, version = parse_sheet_identity("42-1A")
        assert pid == "42"
        assert version == "1A"

    def test_unrecognised_pattern(self):
        pid, version = parse_sheet_identity("Summary")
        assert pid == "Summary"
        assert version == ""


# ---------------------------------------------------------------------------
# ensure_parent_dir
# ---------------------------------------------------------------------------


class TestEnsureParentDir:
    def test_creates_nested_dirs(self, tmp_path):
        target = tmp_path / "a" / "b" / "c" / "file.csv"
        result = ensure_parent_dir(target)
        assert result.parent.exists()
        assert result == target


# ---------------------------------------------------------------------------
# last_populated_header_column
# ---------------------------------------------------------------------------


class TestLastPopulatedHeaderColumn:
    def test_simple(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B", "C"])
        assert last_populated_header_column(ws) == 3

    def test_with_gap(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="A")
        ws.cell(row=1, column=3, value="C")
        assert last_populated_header_column(ws) == 3


# ---------------------------------------------------------------------------
# validate_workbook_schema
# ---------------------------------------------------------------------------


class TestValidateWorkbookSchema:
    def test_valid_workbook(self):
        path = _make_workbook({"1_vA": _VALID_SHEET})
        errors = validate_workbook_schema(path)
        assert errors == []

    def test_no_scoring_sheets(self):
        path = _make_workbook({"Notes": [("Just", "some", "data")]})
        errors = validate_workbook_schema(path)
        assert len(errors) == 1
        assert "Sentence/Stimulus" in errors[0]

    def test_invalid_path(self):
        errors = validate_workbook_schema(Path("/nonexistent/file.xlsx"))
        assert len(errors) == 1
        assert "Cannot open" in errors[0]

    def test_empty_scoring_sheet(self):
        path = _make_workbook({
            "1_vA": [("Sentence", "Stimulus", "Transcription")],  # header only, no data rows
        })
        errors = validate_workbook_schema(path)
        assert any("no sentence rows" in e for e in errors)


# ---------------------------------------------------------------------------
# load_dataset
# ---------------------------------------------------------------------------


class TestLoadDataset:
    def test_loads_valid_workbook(self):
        path = _make_workbook({"1_vA": _VALID_SHEET})
        df = load_dataset(path)
        assert len(df) == 2
        assert list(df.columns) == [
            "sheet_name", "participant_id", "version",
            "sentence_id", "stimulus", "transcription", "human_score",
        ]

    def test_participant_id_extracted(self):
        path = _make_workbook({"42_vB": _VALID_SHEET})
        df = load_dataset(path)
        assert (df["participant_id"] == "42").all()
        assert (df["version"] == "vB").all()

    def test_sorted_by_sentence_id(self):
        rows = [
            ("Sentence", "Stimulus", "Transcription"),
            (3, "C", "c"),
            (1, "A", "a"),
            (2, "B", "b"),
        ]
        path = _make_workbook({"1_vA": rows})
        df = load_dataset(path)
        assert list(df["sentence_id"]) == [1, 2, 3]

    def test_no_scoring_sheet_raises(self):
        path = _make_workbook({"Notes": [("Col1", "Col2")]})
        with pytest.raises(ValueError, match="Sentence/Stimulus"):
            load_dataset(path)

    def test_skips_non_integer_sentence_ids(self):
        rows = [
            ("Sentence", "Stimulus", "Transcription"),
            ("Total", "ignored", "ignored"),
            (1, "Valid", "valid"),
        ]
        path = _make_workbook({"1_vA": rows})
        df = load_dataset(path)
        assert len(df) == 1

    def test_multiple_sheets(self):
        path = _make_workbook({
            "1_vA": _VALID_SHEET,
            "2_vB": _VALID_SHEET,
        })
        df = load_dataset(path)
        assert len(df) == 4


# ---------------------------------------------------------------------------
# write_output_workbook and write_csv_outputs (integration smoke-tests)
# ---------------------------------------------------------------------------


def _make_scored_frame() -> pd.DataFrame:
    return pd.DataFrame([
        {
            "sheet_name": "1_vA",
            "participant_id": "1",
            "version": "vA",
            "sentence_id": 1,
            "stimulus": "El gato corre",
            "transcription": "el gato corre",
            "human_score": 4,
            "auto_score": 4,
            "rationale": "Exact reproduction after normalization",
            "ambiguous_downgraded": False,
            "has_human_score": True,
            "agreement": True,
            "score_diff": 0,
        },
        {
            "sheet_name": "1_vA",
            "participant_id": "1",
            "version": "vA",
            "sentence_id": 2,
            "stimulus": "La niña lee",
            "transcription": "la nina",
            "human_score": 2,
            "auto_score": 2,
            "rationale": "Partial meaning retained",
            "ambiguous_downgraded": False,
            "has_human_score": True,
            "agreement": True,
            "score_diff": 0,
        },
    ])


class TestWriteOutputWorkbook:
    def test_creates_file(self, tmp_path):
        source = _make_workbook({"1_vA": _VALID_SHEET})
        out = tmp_path / "scored.xlsx"
        frame = _make_scored_frame()
        result = write_output_workbook(frame, source_path=source, out_path=out)
        assert result.exists()
        assert result.suffix == ".xlsx"

    def test_summary_sheet_created(self, tmp_path):
        source = _make_workbook({"1_vA": _VALID_SHEET})
        out = tmp_path / "scored.xlsx"
        frame = _make_scored_frame()
        write_output_workbook(frame, source_path=source, out_path=out)
        wb = openpyxl.load_workbook(out)
        assert "AutoEIT_Summary" in wb.sheetnames


class TestWriteCsvOutputs:
    def test_creates_both_files(self, tmp_path):
        frame = _make_scored_frame()
        csv_path = tmp_path / "scores.csv"
        downgrade_path = tmp_path / "downgrades.csv"
        result_csv, result_dl = write_csv_outputs(
            frame, output_csv=csv_path, downgrade_csv=downgrade_path
        )
        assert result_csv.exists()
        assert result_dl.exists()

    def test_csv_has_correct_columns(self, tmp_path):
        frame = _make_scored_frame()
        csv_path = tmp_path / "scores.csv"
        downgrade_path = tmp_path / "downgrades.csv"
        write_csv_outputs(frame, output_csv=csv_path, downgrade_csv=downgrade_path)
        df = pd.read_csv(csv_path)
        assert "auto_score" in df.columns
        assert "rationale" in df.columns
