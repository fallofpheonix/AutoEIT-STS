"""Workbook I/O, schema validation, and formatted output for AutoEIT-STS."""

from __future__ import annotations

import re
from pathlib import Path
from shutil import copyfile

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

# ---------------------------------------------------------------------------
# Cell styling constants
# ---------------------------------------------------------------------------

SCORE_FILLS = {
    4: PatternFill(fill_type="solid", fgColor="C6EFCE"),
    3: PatternFill(fill_type="solid", fgColor="FFEB9C"),
    2: PatternFill(fill_type="solid", fgColor="FFCC99"),
    1: PatternFill(fill_type="solid", fgColor="FFC7CE"),
    0: PatternFill(fill_type="solid", fgColor="FF0000"),
}
HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# ---------------------------------------------------------------------------
# Sheet-identity patterns
# ---------------------------------------------------------------------------

_SHEET_PATTERNS = [
    re.compile(r"^(?P<pid>\d+)_(?P<version>v[AB])$"),
    re.compile(r"^(?P<pid>\d+)-(?P<version>\d[AB])$"),
]


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------


def ensure_parent_dir(path: str | Path) -> Path:
    """Create parent directories for *path* if they do not exist."""
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    return output_path


def last_populated_header_column(worksheet) -> int:
    """Return the 1-based index of the last non-empty header column."""
    for column in range(worksheet.max_column, 0, -1):
        value = worksheet.cell(row=1, column=column).value
        if value is not None and str(value).strip():
            return column
    return 1


def parse_sheet_identity(sheet_name: str) -> tuple[str, str]:
    """Extract (participant_id, version) from a worksheet name.

    Falls back to ``(sheet_name, "")`` if the name does not match any
    recognised pattern.
    """
    for pattern in _SHEET_PATTERNS:
        match = pattern.match(sheet_name)
        if match:
            return match.group("pid"), match.group("version")
    return sheet_name, ""


def _is_scoring_sheet(worksheet) -> bool:
    """Return True if the worksheet has the expected Sentence/Stimulus headers."""
    return (
        worksheet.cell(row=1, column=1).value == "Sentence"
        and worksheet.cell(row=1, column=2).value == "Stimulus"
    )


# ---------------------------------------------------------------------------
# Schema validation
# ---------------------------------------------------------------------------


def validate_workbook_schema(filepath: str | Path) -> list[str]:
    """Validate that a workbook contains at least one well-formed scoring sheet.

    Returns a list of error messages.  An empty list means the workbook
    passed all checks.
    """
    errors: list[str] = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        errors.append(f"Cannot open workbook: {exc}")
        return errors

    scoring_sheets = [name for name in wb.sheetnames if _is_scoring_sheet(wb[name])]
    if not scoring_sheets:
        errors.append(
            "Workbook does not contain any scoring sheets with "
            "Sentence/Stimulus headers in columns A/B of row 1."
        )
        return errors

    for sheet_name in scoring_sheets:
        ws = wb[sheet_name]
        has_data = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if isinstance(row[0], int):
                has_data = True
                if len(row) < 3:
                    errors.append(
                        f"Sheet '{sheet_name}': row {row[0]} has fewer than 3 columns "
                        "(expected Sentence, Stimulus, Transcription)."
                    )
                break
        if not has_data:
            errors.append(f"Sheet '{sheet_name}': no sentence rows found (expected integer in column A).")

    return errors


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------


def load_dataset(filepath: str | Path) -> pd.DataFrame:
    """Parse all scoring sheets in a workbook into a single :class:`~pandas.DataFrame`.

    Raises :exc:`ValueError` if no scoring sheets are found or if the sheets
    contain no sentence rows.
    """
    workbook = openpyxl.load_workbook(filepath, data_only=True)
    records: list[dict[str, object]] = []
    found_sheet = False
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if not _is_scoring_sheet(worksheet):
            continue
        found_sheet = True
        participant_id, version = parse_sheet_identity(sheet_name)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            sentence_id = row[0]
            if not isinstance(sentence_id, int):
                continue
            stimulus = row[1]
            transcription = row[2] if len(row) >= 3 else None
            human_score = row[3] if len(row) >= 4 else None
            records.append(
                {
                    "sheet_name": sheet_name,
                    "participant_id": participant_id,
                    "version": version,
                    "sentence_id": sentence_id,
                    "stimulus": stimulus,
                    "transcription": transcription,
                    "human_score": human_score,
                }
            )
    if not found_sheet:
        raise ValueError(
            "Workbook does not contain any scoring sheets with Sentence/Stimulus headers."
        )
    if not records:
        raise ValueError("Workbook contains scoring sheets but no sentence rows.")
    frame = pd.DataFrame(records)
    return frame.sort_values(["sheet_name", "sentence_id"]).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------


def write_output_workbook(
    frame: pd.DataFrame,
    *,
    source_path: str | Path,
    out_path: str | Path,
) -> Path:
    """Write a scored copy of *source_path* to *out_path*.

    The source workbook is copied verbatim and then annotated with
    AutoEIT_Score / Rationale columns on every scoring sheet, plus an
    AutoEIT_Summary sheet at position 0.
    """
    output = ensure_parent_dir(out_path)
    copyfile(source_path, output)
    workbook = openpyxl.load_workbook(output)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if not _is_scoring_sheet(worksheet):
            continue
        sheet_rows = frame[frame["sheet_name"] == sheet_name]
        lookup = {
            int(row["sentence_id"]): (int(row["auto_score"]), row["rationale"])
            for _, row in sheet_rows.iterrows()
        }
        start_col = last_populated_header_column(worksheet)
        score_col = start_col + 1
        rationale_col = start_col + 2
        worksheet.cell(row=1, column=score_col, value="AutoEIT_Score")
        worksheet.cell(row=1, column=rationale_col, value="Rationale")
        for col in (score_col, rationale_col):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for row_index in range(2, worksheet.max_row + 1):
            sentence_id = worksheet.cell(row=row_index, column=1).value
            if not isinstance(sentence_id, int) or sentence_id not in lookup:
                continue
            auto_score, rationale = lookup[sentence_id]
            worksheet.cell(row=row_index, column=score_col, value=auto_score)
            worksheet.cell(row=row_index, column=score_col).fill = SCORE_FILLS[auto_score]
            worksheet.cell(row=row_index, column=score_col).alignment = Alignment(horizontal="center")
            worksheet.cell(row=row_index, column=rationale_col, value=rationale)
            worksheet.cell(row=row_index, column=rationale_col).alignment = Alignment(wrap_text=True)

    if "AutoEIT_Summary" in workbook.sheetnames:
        del workbook["AutoEIT_Summary"]
    summary = workbook.create_sheet("AutoEIT_Summary", 0)
    headers = [
        "Sheet", "Participant", "Version", "Sentence", "Stimulus",
        "Transcription", "Human Score", "AutoEIT Score", "Score Diff", "Rationale",
    ]
    widths = [18, 14, 10, 10, 50, 55, 13, 13, 11, 60]
    for index, header in enumerate(headers, start=1):
        cell = summary.cell(row=1, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        summary.column_dimensions[get_column_letter(index)].width = widths[index - 1]
    for row_index, (_, row) in enumerate(frame.iterrows(), start=2):
        diff: int | str = ""
        if pd.notna(row.get("human_score")):
            diff = int(row["auto_score"] - row["human_score"])
        values = [
            row["sheet_name"],
            row["participant_id"],
            row["version"],
            int(row["sentence_id"]),
            row["stimulus"],
            row["transcription"],
            int(row["human_score"]) if pd.notna(row["human_score"]) else "",
            int(row["auto_score"]),
            diff,
            row["rationale"],
        ]
        for col_index, value in enumerate(values, start=1):
            cell = summary.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(wrap_text=True)
            if col_index == 8:
                cell.fill = SCORE_FILLS[int(row["auto_score"])]
                cell.alignment = Alignment(horizontal="center")
    summary.freeze_panes = "A2"
    workbook.save(output)
    return output


def write_csv_outputs(
    frame: pd.DataFrame,
    *,
    output_csv: str | Path,
    downgrade_csv: str | Path,
) -> tuple[Path, Path]:
    """Write the main CSV output and the ambiguous-downgrade log."""
    csv_path = ensure_parent_dir(output_csv)
    downgrade_path = ensure_parent_dir(downgrade_csv)
    frame[
        [
            "sheet_name", "participant_id", "version", "sentence_id",
            "stimulus", "transcription", "human_score", "auto_score",
            "ambiguous_downgraded", "rationale",
        ]
    ].to_csv(csv_path, index=False)
    frame[frame["ambiguous_downgraded"]][
        [
            "sheet_name", "participant_id", "version", "sentence_id",
            "stimulus", "transcription", "human_score", "auto_score", "rationale",
        ]
    ].to_csv(downgrade_path, index=False)
    return csv_path, downgrade_path
