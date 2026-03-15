"""Deterministic Task II scorer for AutoEIT."""

from __future__ import annotations

from pathlib import Path
from shutil import copyfile
import re
import unicodedata

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

AMBIGUOUS_TO_2 = True
AMBIGUOUS_EDIT_MIN = 2
AMBIGUOUS_EDIT_MAX = 4
AMBIGUOUS_OVERLAP_MIN = 0.60
AMBIGUOUS_OVERLAP_MAX = 0.78

SCORE_FILLS = {
    4: PatternFill(fill_type="solid", fgColor="C6EFCE"),
    3: PatternFill(fill_type="solid", fgColor="FFEB9C"),
    2: PatternFill(fill_type="solid", fgColor="FFCC99"),
    1: PatternFill(fill_type="solid", fgColor="FFC7CE"),
    0: PatternFill(fill_type="solid", fgColor="FF0000"),
}
HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)

FUNCTION_WORDS = {
    "a", "al", "como", "con", "cuya", "cuyo", "de", "del", "el", "ella",
    "ellos", "ellas", "en", "era", "eran", "es", "ese", "esa", "esos", "esas",
    "esta", "estan", "estar", "este", "estos", "fue", "fueron", "ha", "haber",
    "han", "hay", "he", "hemos", "la", "las", "le", "les", "lo", "los", "mas",
    "me", "mi", "mis", "muy", "ni", "no", "nos", "nosotros", "o", "para",
    "pero", "por", "que", "quien", "se", "ser", "si", "sin", "sobre", "son",
    "su", "sus", "te", "tener", "tu", "tus", "un", "una", "unos", "unas",
    "usted", "ustedes", "y", "ya", "yo",
}

SER_FORMS = {"es", "son", "era", "eran", "fue", "fueron", "sea", "sean", "ser", "soy", "somos"}
ESTAR_FORMS = {"esta", "estan", "estaba", "estaban", "estuvo", "estuvieron", "estar", "estoy", "estamos"}
BRACKET_CONTENT = re.compile(r"\[.*?\]")
PUNCTUATION = re.compile(r"[¿?¡!,.:;«»\"'\`´]")
WHITESPACE = re.compile(r"\s+")
WORD_COUNT_HINT = re.compile(r"\(\d+\)\s*$")


def ensure_parent_dir(path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    return output_path


def last_populated_header_column(worksheet) -> int:
    for column in range(worksheet.max_column, 0, -1):
        value = worksheet.cell(row=1, column=column).value
        if value is not None and str(value).strip():
            return column
    return 1


def strip_accents(text: str) -> str:
    return "".join(
        char for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    )


def normalize_stimulus_text(text: str) -> str:
    if not isinstance(text, str):
        return ""
    normalized = WORD_COUNT_HINT.sub("", text)
    normalized = BRACKET_CONTENT.sub(" ", normalized)
    normalized = PUNCTUATION.sub("", normalized)
    normalized = strip_accents(normalized.lower())
    return WHITESPACE.sub(" ", normalized).strip()


def normalize_transcription_text(text: str) -> str:
    if not isinstance(text, str):
        return ""
    normalized = BRACKET_CONTENT.sub(" ", text)
    normalized = PUNCTUATION.sub("", normalized)
    normalized = strip_accents(normalized.lower())
    return WHITESPACE.sub(" ", normalized).strip()


def clean(text: str, *, is_stimulus: bool = False) -> str:
    if is_stimulus:
        return normalize_stimulus_text(text)
    return normalize_transcription_text(text)


def tokenise(text: str, *, is_stimulus: bool = False) -> list[str]:
    return clean(text, is_stimulus=is_stimulus).split()


def token_edit_distance(seq_a: list[str], seq_b: list[str]) -> int:
    rows = len(seq_a) + 1
    cols = len(seq_b) + 1
    dp = list(range(cols))
    for row in range(1, rows):
        previous = dp[:]
        dp[0] = row
        for col in range(1, cols):
            if seq_a[row - 1] == seq_b[col - 1]:
                dp[col] = previous[col - 1]
            else:
                dp[col] = 1 + min(previous[col], dp[col - 1], previous[col - 1])
    return dp[-1]


def content_words(tokens: list[str]) -> list[str]:
    return [token for token in tokens if token not in FUNCTION_WORDS and len(token) > 1]


def content_overlap_ratio(target_tokens: list[str], learner_tokens: list[str]) -> float:
    target_content = content_words(target_tokens)
    learner_content = set(content_words(learner_tokens))
    if not target_content:
        return 1.0
    matched = sum(1 for token in target_content if token in learner_content)
    return matched / len(target_content)


def is_no_response(raw: str) -> bool:
    if not isinstance(raw, str) or not raw.strip():
        return True
    cleaned = BRACKET_CONTENT.sub("", raw).strip()
    cleaned = PUNCTUATION.sub("", cleaned).strip()
    return cleaned == ""


def is_wrong_language(raw: str) -> bool:
    if not isinstance(raw, str):
        return False
    lowered = raw.lower()
    return "en ingles" in lowered or "en inglés" in lowered


def only_ser_estar_swap(target_tokens: list[str], learner_tokens: list[str]) -> bool:
    if len(target_tokens) != len(learner_tokens):
        return False
    differences = [(left, right) for left, right in zip(target_tokens, learner_tokens) if left != right]
    if len(differences) != 1:
        return False
    left, right = differences[0]
    return (left in SER_FORMS and right in ESTAR_FORMS) or (left in ESTAR_FORMS and right in SER_FORMS)


def char_similarity(left: str, right: str) -> float:
    if left == right:
        return 1.0
    max_length = max(len(left), len(right), 1)
    return 1 - token_edit_distance(list(left), list(right)) / max_length


def morph_close(left: str, right: str, *, threshold: float = 0.72) -> bool:
    return char_similarity(left, right) >= threshold


def is_ambiguous_meaning_case(edit_distance: int, overlap: float) -> bool:
    return (
        AMBIGUOUS_TO_2
        and AMBIGUOUS_EDIT_MIN <= edit_distance <= AMBIGUOUS_EDIT_MAX
        and AMBIGUOUS_OVERLAP_MIN <= overlap <= AMBIGUOUS_OVERLAP_MAX
    )


def score_utterance(
    target_raw: str,
    learner_raw: str,
    *,
    return_meta: bool = False,
) -> tuple[int, str] | tuple[int, str, bool]:
    ambiguous_downgraded = False

    def _ret(score: int, rationale: str):
        if return_meta:
            return score, rationale, ambiguous_downgraded
        return score, rationale

    if is_wrong_language(learner_raw):
        return _ret(0, "Response in wrong language [en ingles]")
    if is_no_response(learner_raw):
        return _ret(0, "No response or empty transcription")

    target_tokens = tokenise(target_raw, is_stimulus=True)
    learner_tokens = tokenise(learner_raw)
    if not target_tokens:
        return _ret(0, "Target sentence is empty")

    edit_distance = token_edit_distance(target_tokens, learner_tokens)
    overlap = content_overlap_ratio(target_tokens, learner_tokens)
    learner_count = len(learner_tokens)
    target_count = len(target_tokens)

    if edit_distance == 0:
        return _ret(4, "Exact reproduction after normalization")
    if edit_distance == 1 and overlap >= 0.67:
        return _ret(4, f"1-word deviation, meaning preserved (overlap={overlap:.2f})")
    if edit_distance == 2 and overlap >= 0.80:
        return _ret(4, f"2-word deviation consistent with disfluency artifact (overlap={overlap:.2f})")
    if edit_distance == 3 and overlap >= 0.85:
        return _ret(4, f"3-word deviation with near-complete content retention (overlap={overlap:.2f})")

    if only_ser_estar_swap(target_tokens, learner_tokens):
        return _ret(3, "ser/estar substitution with preserved meaning")
    if edit_distance == 1 and overlap >= 0.50:
        differences = [(left, right) for left, right in zip(target_tokens, learner_tokens) if left != right]
        if differences:
            left, right = differences[0]
            if morph_close(left, right):
                return _ret(3, f"Morphological variant '{left}' -> '{right}' with preserved meaning")
        return _ret(3, f"Single-word deviation with preserved meaning (overlap={overlap:.2f})")
    if edit_distance == 2 and overlap >= 0.70 and learner_count >= target_count - 1:
        return _ret(3, f"Near-accurate response with high content retention (overlap={overlap:.2f})")
    if edit_distance == 2 and overlap >= 0.50 and learner_count >= target_count - 1:
        if is_ambiguous_meaning_case(edit_distance, overlap):
            ambiguous_downgraded = True
            return _ret(2, f"Ambiguous meaning boundary downgraded to 2 (edit={edit_distance}, overlap={overlap:.2f})")
        return _ret(3, f"Near-accurate response with limited omission/substitution (overlap={overlap:.2f})")
    if edit_distance in (2, 3) and overlap >= 0.65:
        if is_ambiguous_meaning_case(edit_distance, overlap):
            ambiguous_downgraded = True
            return _ret(2, f"Ambiguous meaning boundary downgraded to 2 (edit={edit_distance}, overlap={overlap:.2f})")
        return _ret(3, f"Minor structural deviation with preserved meaning (edit={edit_distance}, overlap={overlap:.2f})")
    if edit_distance <= 4 and overlap >= 0.75:
        if is_ambiguous_meaning_case(edit_distance, overlap):
            ambiguous_downgraded = True
            return _ret(2, f"Ambiguous meaning boundary downgraded to 2 (edit={edit_distance}, overlap={overlap:.2f})")
        return _ret(3, f"High content retention despite structural deviation (edit={edit_distance}, overlap={overlap:.2f})")

    if learner_count < 3 and edit_distance >= 4 and overlap > 0:
        return _ret(1, f"Fragmentary response only (edit={edit_distance}, overlap={overlap:.2f})")

    if overlap >= 0.35 and learner_count >= max(2, int(target_count * 0.30)):
        return _ret(2, f"Partial meaning retained (edit={edit_distance}, overlap={overlap:.2f})")
    if overlap >= 0.25 and learner_count >= 2:
        return _ret(2, f"Partial response with some retained content (overlap={overlap:.2f})")
    if overlap > 0:
        return _ret(1, f"Minimal fragmentary meaning only (overlap={overlap:.2f})")
    return _ret(0, f"No meaning preserved (edit={edit_distance}, overlap={overlap:.2f})")


def parse_sheet_identity(sheet_name: str) -> tuple[str, str]:
    patterns = [
        re.compile(r"^(?P<pid>\d+)_(?P<version>v[AB])$"),
        re.compile(r"^(?P<pid>\d+)-(?P<version>\d[AB])$"),
    ]
    for pattern in patterns:
        match = pattern.match(sheet_name)
        if match:
            return match.group("pid"), match.group("version")
    return sheet_name, ""


def _is_scoring_sheet(worksheet) -> bool:
    return worksheet.cell(row=1, column=1).value == "Sentence" and worksheet.cell(row=1, column=2).value == "Stimulus"


def load_dataset(filepath: str | Path) -> pd.DataFrame:
    workbook = openpyxl.load_workbook(filepath, data_only=True)
    records: list[dict[str, object]] = []
    found_sheet = False
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if not _is_scoring_sheet(worksheet):
            continue
        found_sheet = True
        participant_id, version = parse_sheet_identity(sheet_name)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            sentence_id = row[0]
            if not isinstance(sentence_id, int):
                continue
            stimulus = row[1]
            transcription = row[2] if len(row) >= 3 else None
            human_score = row[3] if len(row) >= 4 else None
            records.append(
                {
                    "sheet_name": sheet_name,
                    "participant_id": participant_id,
                    "version": version,
                    "sentence_id": sentence_id,
                    "stimulus": stimulus,
                    "transcription": transcription,
                    "human_score": human_score,
                }
            )
    if not found_sheet:
        raise ValueError("Workbook does not contain any scoring sheets with Sentence/Stimulus headers.")
    if not records:
        raise ValueError("Workbook contains scoring sheets but no sentence rows.")
    frame = pd.DataFrame(records)
    return frame.sort_values(["sheet_name", "sentence_id"]).reset_index(drop=True)


def run_scoring_pipeline(filepath: str | Path) -> pd.DataFrame:
    frame = load_dataset(filepath)
    scored = frame.apply(
        lambda row: score_utterance(row["stimulus"], row["transcription"], return_meta=True),
        axis=1,
    )
    frame["auto_score"] = scored.apply(lambda result: result[0])
    frame["rationale"] = scored.apply(lambda result: result[1])
    frame["ambiguous_downgraded"] = scored.apply(lambda result: result[2])
    frame["has_human_score"] = frame["human_score"].notna()
    frame.loc[frame["has_human_score"], "agreement"] = (
        frame.loc[frame["has_human_score"], "auto_score"]
        == frame.loc[frame["has_human_score"], "human_score"]
    )
    frame.loc[frame["has_human_score"], "score_diff"] = (
        frame.loc[frame["has_human_score"], "auto_score"]
        - frame.loc[frame["has_human_score"], "human_score"]
    )
    return frame


def compute_metrics(frame: pd.DataFrame) -> dict[str, object]:
    rated = frame[frame["has_human_score"]].copy()
    if rated.empty:
        return {}
    exact = float(rated["agreement"].sum()) / len(rated) * 100
    within_one = float((rated["score_diff"].abs() <= 1).sum()) / len(rated) * 100
    auto_totals = rated.groupby(["sheet_name"])["auto_score"].sum()
    human_totals = rated.groupby(["sheet_name"])["human_score"].sum()
    deviations = (auto_totals - human_totals).abs()
    return {
        "n_rated_utterances": len(rated),
        "exact_agreement_rate_pct": round(exact, 2),
        "within1_agreement_pct": round(within_one, 2),
        "mean_participant_deviation": round(float(deviations.mean()), 2),
        "max_participant_deviation": int(deviations.max()),
        "pct_participants_within10": round(float((deviations <= 10).mean() * 100), 2),
        "n_participants": int(len(deviations)),
        "n_ambiguous_downgraded": int(frame["ambiguous_downgraded"].sum()),
        "auto_score_distribution": rated["auto_score"].value_counts().sort_index().to_dict(),
        "human_score_distribution": rated["human_score"].value_counts().sort_index().to_dict(),
        "confusion_summary": pd.crosstab(
            rated["human_score"],
            rated["auto_score"],
            rownames=["Human"],
            colnames=["Auto"],
        ),
    }


def write_output_workbook(
    frame: pd.DataFrame,
    *,
    source_path: str | Path,
    out_path: str | Path,
) -> Path:
    output = ensure_parent_dir(out_path)
    copyfile(source_path, output)
    workbook = openpyxl.load_workbook(output)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if not _is_scoring_sheet(worksheet):
            continue
        sheet_rows = frame[frame["sheet_name"] == sheet_name]
        lookup = {
            int(row["sentence_id"]): (int(row["auto_score"]), row["rationale"])
            for _, row in sheet_rows.iterrows()
        }
        start_col = last_populated_header_column(worksheet)
        score_col = start_col + 1
        rationale_col = start_col + 2
        worksheet.cell(row=1, column=score_col, value="AutoEIT_Score")
        worksheet.cell(row=1, column=rationale_col, value="Rationale")
        for col in (score_col, rationale_col):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for row_index in range(2, worksheet.max_row + 1):
            sentence_id = worksheet.cell(row=row_index, column=1).value
            if not isinstance(sentence_id, int) or sentence_id not in lookup:
                continue
            auto_score, rationale = lookup[sentence_id]
            worksheet.cell(row=row_index, column=score_col, value=auto_score)
            worksheet.cell(row=row_index, column=score_col).fill = SCORE_FILLS[auto_score]
            worksheet.cell(row=row_index, column=score_col).alignment = Alignment(horizontal="center")
            worksheet.cell(row=row_index, column=rationale_col, value=rationale)
            worksheet.cell(row=row_index, column=rationale_col).alignment = Alignment(wrap_text=True)

    if "AutoEIT_Summary" in workbook.sheetnames:
        del workbook["AutoEIT_Summary"]
    summary = workbook.create_sheet("AutoEIT_Summary", 0)
    headers = [
        "Sheet",
        "Participant",
        "Version",
        "Sentence",
        "Stimulus",
        "Transcription",
        "Human Score",
        "AutoEIT Score",
        "Score Diff",
        "Rationale",
    ]
    widths = [18, 14, 10, 10, 50, 55, 13, 13, 11, 60]
    for index, header in enumerate(headers, start=1):
        cell = summary.cell(row=1, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        summary.column_dimensions[get_column_letter(index)].width = widths[index - 1]
    for row_index, (_, row) in enumerate(frame.iterrows(), start=2):
        diff = ""
        if pd.notna(row.get("human_score")):
            diff = int(row["auto_score"] - row["human_score"])
        values = [
            row["sheet_name"],
            row["participant_id"],
            row["version"],
            int(row["sentence_id"]),
            row["stimulus"],
            row["transcription"],
            int(row["human_score"]) if pd.notna(row["human_score"]) else "",
            int(row["auto_score"]),
            diff,
            row["rationale"],
        ]
        for col_index, value in enumerate(values, start=1):
            cell = summary.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(wrap_text=True)
            if col_index == 8:
                cell.fill = SCORE_FILLS[int(row["auto_score"])]
                cell.alignment = Alignment(horizontal="center")
    summary.freeze_panes = "A2"
    workbook.save(output)
    return output


def write_csv_outputs(
    frame: pd.DataFrame,
    *,
    output_csv: str | Path,
    downgrade_csv: str | Path,
) -> tuple[Path, Path]:
    csv_path = ensure_parent_dir(output_csv)
    downgrade_path = ensure_parent_dir(downgrade_csv)
    frame[
        [
            "sheet_name",
            "participant_id",
            "version",
            "sentence_id",
            "stimulus",
            "transcription",
            "human_score",
            "auto_score",
            "ambiguous_downgraded",
            "rationale",
        ]
    ].to_csv(csv_path, index=False)
    frame[frame["ambiguous_downgraded"]][
        [
            "sheet_name",
            "participant_id",
            "version",
            "sentence_id",
            "stimulus",
            "transcription",
            "human_score",
            "auto_score",
            "rationale",
        ]
    ].to_csv(downgrade_path, index=False)
    return csv_path, downgrade_path


def score_workbook(
    *,
    source_path: str | Path,
    output_xlsx: str | Path,
    output_csv: str | Path,
    downgrade_csv: str | Path | None = None,
) -> tuple[pd.DataFrame, dict[str, object], Path, Path, Path]:
    frame = run_scoring_pipeline(source_path)
    metrics = compute_metrics(frame)
    workbook_path = write_output_workbook(frame, source_path=source_path, out_path=output_xlsx)
    downgrade_path = downgrade_csv or Path(output_csv).with_name("AutoEIT_ambiguous_downgrades.csv")
    csv_path, downgrade_path = write_csv_outputs(
        frame,
        output_csv=output_csv,
        downgrade_csv=downgrade_path,
    )
    return frame, metrics, workbook_path, csv_path, downgrade_path
